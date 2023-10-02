import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
import openpyxl

# Carregue o arquivo Excel existente
workbook = openpyxl.load_workbook('teste.xlsx')

# Selecione a planilha em que as chaves estão localizadas
sheet_chaves = workbook.active  # Use a planilha ativa ou selecione uma planilha específica

# Inicialize o número da linha
linha_atual = 1

# Percorra as linhas da coluna de chaves e obtenha as datas correspondentes
for chave in sheet_chaves.iter_rows(min_col=1, max_col=1, values_only=True):
    chave = chave[0]

    # Crie a URL com base na chave
    url = f'https://exemplo.com.br/chave={chave}'

    # Faça a solicitação HTTP para obter o conteúdo da página
    response = requests.get(url)

    # Verifique se a solicitação foi bem-sucedida
    if response.status_code == 200:
        # Parseie o conteúdo HTML da página com BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')

        # Encontre a tag <p> que contém a data/hora
        data_tag = soup.find('p', text=re.compile(r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}-\d{2}:\d{2}'))

        # Extraia o texto da tag <p>
        data_text = data_tag.get_text(strip=True)

        # Converta a string em um objeto datetime
        data_hora_emissao = datetime.strptime(data_text, '%d/%m/%Y %H:%M:%S%z')
        so_a_data = data_hora_emissao.date()

        # Adicione a data à coluna 3 da mesma linha
        sheet_chaves.cell(row=linha_atual, column=3, value=so_a_data)

        # Atualize o número da linha
        linha_atual += 1

    else:
        print(f"Falha ao acessar a página para a chave {chave}. Código de status:", response.status_code)

# Salve as alterações no arquivo Excel
workbook.save('teste.xlsx')

